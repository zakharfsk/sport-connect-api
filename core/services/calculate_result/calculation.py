"""
Функція має повертати список словників в вигляді:

[
  {
    "id": "",
    "first_name": "",
    "last_name": "",
    "gender": ""
    "age": "",
    "school": "",
    "classrom": "",
    "standards": {
      "Назва_екзамену": "результат дитини"
    }
  }
]
"""
import logging
from pprint import pprint

from openpyxl import load_workbook
from django.conf import settings
from django.core.files.storage import default_storage

from core.models import AverageValuesStandards, Sport, WeightingFactors, SportStandard

logger = logging.getLogger(__name__)


def get_data_from_excel(file_path: str) -> list:
    """
    This method provides functionality to extract data from an Excel file.

    Parameters:
    - file_path (str): The path to the Excel file from which to extract the data.

    Returns:
    - list: A list of dictionaries, where each dictionary represents a row of data from the Excel file.

    The method works by loading the Excel file using the `load_workbook` function from the `openpyxl` library. It then retrieves the active worksheet using the `active` attribute of the
    * workbook. The method iterates over each row of the worksheet starting from the second row (index 2) and builds a dictionary for each row. The keys of the dictionary are taken from
    * a pre-defined list called `settings.LIST_TOPICS`. The values of the dictionary are extracted from the corresponding cells in the row.

    If a value in a cell is not None, it is added to the dictionary as a key-value pair. If the index of the column is greater than 7, the value is added to a nested dictionary under the
    * "standards" key. Otherwise, it is added directly to the top-level dictionary.

    After processing all rows in the worksheet, the method calculates two additional values and adds them to the "standards" dictionary. The first calculated value is the "Вагово-ростов
    *ий індекс (індекс маси тіла), маса (гр), зріст (см)" which is the ratio of the "Зріст, см" to the "Маса, гр". The second calculated value is the "Індекс розвитку мускулатури (перим
    *етр плеча напруженого/периметр плеча розслабленого)" which is the ratio of the "Периметр плеча напруженого, см" to the "Периметр плеча розслабленого, см".

    Finally, the method deletes the Excel file using the `default_storage.delete` method and returns the list of dictionaries representing the extracted data.

    """
    wb = load_workbook(file_path)
    ws = wb.active
    data = []

    for j in range(2, ws.max_row + 1):
        d = {}
        for i, key in enumerate(settings.LIST_TOPICS, start=1):
            val = ws.cell(row=j, column=i).value
            if val is not None:
                if i > 7:
                    d.setdefault("standards", dict())
                    d["standards"][key] = val
                else:
                    d[key] = val
        d["standards"][
            "Ваго-ростовий індекс (індекс маси тіла)"
        ] = d["standards"]['Маса, гр'] / d["standards"]['Зріст, см']

        d["standards"]["Індекс розвитку мускулатури (периметр плеча напруженого/периметр плеча розслабленого)"] = \
            ((d["standards"]['Периметр плеча напруженого, см'] - d["standards"]['Периметр плеча розслабленого, см']) /
             d["standards"]['Периметр плеча розслабленого, см'] * 100)

        d["standards"]["Співвідношення розмаху рук до довжини тіла стоячи, см"] = \
            d["standards"]["Ширина рук, см"] - d["standards"]['Зріст, см']

        d["standards"]["Викрут мірної лінійки, см"] = \
            d["standards"]["Викрут мірної лінійки, см"] - d["standards"]["Ширина плечей, см"]

        d.get("standards", {}).pop('Маса, гр')
        d.get("standards", {}).pop('Периметр плеча розслабленого, см')
        d.get("standards", {}).pop('Периметр плеча напруженого, см')
        d.get("standards", {}).pop('Ширина плечей, см')
        d.get("standards", {}).pop('Ширина рук, см')
        data.append(d)

    default_storage.delete(file_path)

    return data


def calculate_standards_result(file_result: list):
    """
    Calculate the standards result based on the given file_result.

    :param file_result: A list of dictionaries representing the file result. Each dictionary contains the following keys:
        - 'id': The id of the result.
        - 'standards': A dictionary of user standards as keys and their respective values.

    :return: A list of dictionaries representing the calculated standards result.
        Each dictionary contains the following keys:
        - 'id': The id of the result.
        - User standards as keys and their respective calculated results.

    """
    # print(file_result)
    data = []
    for res in file_result:
        d = {"id": res["id"]}
        for user_standards, value in res['standards'].items():
            try:
                # print(f"standard name:{user_standards}")
                # print(f"value:{value}")
                average = AverageValuesStandards.objects.get(
                    standard__name=user_standards,
                    children_age=res['Вік'],
                    children_gender=res['Стать']
                )

                result = 50 + 10 * ((value - average.average_value) / average.sigma)
                d[user_standards] = result
                d[user_standards + "_avg"] = average.average_value
                d[user_standards + "_sigma"] = average.sigma
            except AverageValuesStandards.DoesNotExist:
                logger.warning(f"calculate_standards_result {user_standards} not found")

        data.append(d)
    # print("result:")
    # print(data[0])
    return data


def calculate_sports_aptitude(standards_result: list):
    """
    Calculate the sports aptitude for each standard result.

    Parameters:
    - standards_result: A list of dictionaries containing the results for each standard. Each dictionary should have the following keys:
      - "id": The ID of the standard result.
      - Any other key-value pairs representing the results of individual standards.

    Returns:
    - A list of dictionaries containing the calculated sports aptitude for each standard result. Each dictionary will have the following keys:
      - "id": The ID of the standard result.
      - Any other key-value pairs representing the calculated sports aptitude for each sport.

    Example Usage:
    standards_result = [
      {"id": 1, "running": 10, "jumping": 8, "throwing": 7},
      {"id": 2, "running": 9, "jumping": 6, "throwing": 8},
      {"id": 3, "running": 7, "jumping": 9, "throwing": 6},
    ]
    aptitude_results = calculate_sports_aptitude(standards_result)
    print(aptitude_results)
    # Output: [{"id": 1, "football": 70, "tennis": 65, "swimming": 68},
    #          {"id": 2, "football": 64, "tennis": 63, "swimming": 62},
    #          {"id": 3, "football": 65, "tennis": 66, "swimming": 61}]

    Note:
    - This method assumes that there is a Sport model and a WeightingFactors model in the system.
    - The Sport model should have a "name" field representing the name of the sport.
    - The WeightingFactors model should have foreign keys to the Sport model and the AverageValueStandard model.
    - The AverageValueStandard model should have a "name_standard" field representing the name of the standard.
    - The WeightingFactors model also should have a "weighting_factor" field representing the factor to be used for calculating the sports aptitude.
    """
    result = []

    for st_res in standards_result:

        result_dict = {"id": st_res["id"], "sport_results": []}

        for sport in Sport.objects.all():
            weight_factors = dict(
                WeightingFactors.objects.filter(sport=sport.pk)
                .values_list("sport_standard__name", "weighting_factor")
            )

            #str_weights=""
            res = 0
            for key, value in st_res.items():
                if key in weight_factors:
                    res += value * weight_factors[key]
                    #str_weights += str(key) + "=" + str(weight_factors[key])+","

            result_dict["sport_results"].append({sport.name: res})
        result.append(result_dict)
    return result
